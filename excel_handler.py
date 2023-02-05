import sys
import openpyxl
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from dateutil.relativedelta import relativedelta

class ExcelHandler:

    last_row = 0

    def initialize_output_file(self, outputfile):
        print(sys._getframe().f_code.co_name, outputfile)
        wb = openpyxl.Workbook() # workbook
        ws = wb.active # worksheet
        ws['A1'] = "VES rate by USD"
        ws['A2'] = "Date"
        ws['B2'] = "Rate"
        self.last_row = 2
        wb.save(outputfile)
        wb.close()

    def process_file(self, inputfile, start_date, end_date):
        print(sys._getframe().f_code.co_name, inputfile)
        print(f'start: {start_date.strftime("%d %m %Y")} - end: {end_date.strftime("%d %m %Y")}')
        wb = openpyxl.load_workbook(inputfile, read_only=True) # workbook
        results = list()
        next_date = start_date
        while next_date != end_date:
            wsname = next_date.strftime("%d%m%Y")
            if wsname in wb.sheetnames:
                ws = wb[wsname]
                row = self.__get_currency_row(ws)
                value = self.__get_currency(ws, row).value
                results.append(tuple([next_date, value]))
            else:
                results.append(tuple([next_date, None]))
            next_date = next_date + relativedelta(days=1)
        return results

    def append_output_to_file(self, output, outputfile):
        print(sys._getframe().f_code.co_name)
        wb = openpyxl.load_workbook(outputfile) # workbook
        ws = wb.active # worksheet
        old_last_row = self.last_row
        for index, result in enumerate(output):
            row = old_last_row + index + 1
            datecell = ws.cell(row, 1, value=result[0])
            datecell.number_format = "DD/MM/YYYY"
            ws.cell(row, 2, value=result[1])
            self.last_row = row
        self.__adjust_column_width(ws)
        wb.save(outputfile)
        wb.close()

    def __get_currency_row(self, worksheet):
        col = 2
        for row in range(11, 33):
            cell = worksheet.cell(row, col)
            if cell.value == "USD":
                return row
        return None

    def __get_currency(self, worksheet, row):
        col = 7
        return worksheet.cell(row, col)

    def __adjust_column_width(self, worksheet):
        dim_holder = DimensionHolder(worksheet=worksheet)
        for col in range(worksheet.min_column, worksheet.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(worksheet, min=col, max=col, width=12)
        worksheet.column_dimensions = dim_holder